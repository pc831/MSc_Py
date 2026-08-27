"""Build the aluminium asset-transition milestone workbook from MPP and IAI published data.

Mirrors derive_milestones.py, but formula-driven end to end so the derivation can be walked
through in Excel. Only raw source tables and the emission/fuel constants are written as
values; every classification, aggregation and milestone test is a live formula.

Sheets, in reading order:
  Read Me               what this is, sources, method, how to navigate
  Inputs                the five parameters, as named cells every formula points at
  Abatement Test        which technologies are unabated, and why
  Published Production  MPP production, with technology strings split into components
  Milestone Digester    the two tests for refinery boilers
  Milestone Calciner    the two tests for refinery calciners
  Milestone Carbon Anode  the two tests for smelter anodes
  Auxiliary IAI         the auxiliary milestone, on the IAI intensity curve
  Milestone Summary     the specification sheet table, plus checks against the script
"""

from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

DATA_DIR = Path(".")
MPP = DATA_DIR / "mpp-shared-code" / "aluminium" / "data" / "lc"
OUT = Path("/Users/parkercaswell/Desktop/Aluminium Pathway Derivation/"
           "Aluminium Milestone Pathway.xlsx")

PUBLISHED = {"China": "mpp_aluminium_net_zero_outputs (1).xlsx",
             "RoW": "mpp_aluminium_net_zero_outputs (2).xlsx"}
SCENARIO = "1.5DS"
REGIONS = ["China", "RoW"]

# Emission factor provenance. Middle East 2030 is the only MPP region carrying all seven
# boiler types, and the component ratios are identical in every region and year, so one
# region-year is enough to classify. Components are isolated by pairing each boiler with
# the zero-emitting Elec-Calciner, and each calciner with the zero-emitting Elec-Boiler.
EF_REGION, EF_YEAR = "Middle East", 2030
ANODE_EF_REGION, ANODE_EF_YEAR = "China - North", 2030

# The one technology whose rule-based classification is overridden, with its reason.
OVERRIDES = {"CST-Fossil-Boiler": (
    "Abated",
    "Concentrated solar thermal. Fails the 10% intensity test at 26%, but the residual is a "
    "backup rather than a primary fuel and is treated as switchable to biogas or hydrogen.")}

# MPP specifies HH+CCS as 90% capture of smelter process CO2, not PFCs or anode production
# emissions (Technical Appendix p.11). No other technology in the set carries CCS.
CAPTURE_RATES = {"Carbon Anode+CCS": 0.90}

MILESTONES = [
    {"sheet": "Milestone Digester", "label": "unabated fossil digester",
     "component_col": "D", "class_col": "J", "unit": "Mt alumina",
     "technologies": ["Coal-Boiler", "Oil-Boiler", "Gas-Boiler"]},
    {"sheet": "Milestone Calciner", "label": "unabated fossil calciner",
     "component_col": "E", "class_col": "K", "unit": "Mt alumina",
     "technologies": ["Gas-Calciner", "Oil-Calciner"]},
    {"sheet": "Milestone Carbon Anode", "label": "unabated carbon anode",
     "component_col": "F", "class_col": "L", "unit": "Mt aluminium",
     "technologies": ["Carbon Anode"]},
]

# IAI 1.5DS value chain steps that make up auxiliary, and the throughput they are measured
# against. Primary casting is not separable - IAI folds it into Primary Aluminium.
IAI_STEPS = ["Recycled Aluminium", "Internal Scrap/Fabrication Scrap", "Semis Process"]
IAI_DENOMINATOR = "Semis Shipments"

HEAD = Font(bold=True)
TITLE = Font(bold=True, size=14)
SUBHEAD = Font(bold=True, size=12)


def load_repo(path):
    df = pd.read_csv(path)
    df.columns = [str(c).strip().lstrip("﻿") for c in df.columns]
    return df


# ============================================================== source data
published = []
for region, filename in PUBLISHED.items():
    df = pd.read_excel(DATA_DIR / filename, sheet_name="Annual_production_volume_Mt_df")
    df = df[df["scenario"] == SCENARIO].copy()
    df["Region"] = region
    published.append(df)
published = pd.concat(published)

production = published[["Region", "plant_type", "technology", "year", "value"]].copy()
production.columns = ["Region", "Plant Type", "Technology", "Year", "Production Mt"]
production = production.sort_values(
    ["Region", "Plant Type", "Technology", "Year"]).reset_index(drop=True)
YEARS = sorted(production["Year"].unique())

# Refinery emission factors. Boiler component = scope 1 of "<boiler> + Elec-Calciner";
# calciner component = scope 1 of "Elec-Boiler + <calciner>".
ref_ef = load_repo(MPP / "def_refineries" / "intermediate" / "emissions.csv")
ref_ef = ref_ef[(ref_ef["region"] == EF_REGION) & (ref_ef["year"] == EF_YEAR)]
ref_ef = ref_ef.set_index("technology")["co2_scope1"]

# Smelter anode factors, read on "+ Grid" so captive generation contributes zero.
sm_ef = load_repo(MPP / "def" / "intermediate" / "emissions.csv")
sm_ef = sm_ef[(sm_ef["region"] == ANODE_EF_REGION) & (sm_ef["year"] == ANODE_EF_YEAR)]
sm_ef = sm_ef.set_index("technology")["co2_scope1"]

boilers = sorted({t.split(" + ")[0] for t in production.loc[
    production["Plant Type"] == "Refinery", "Technology"]})
calciners = sorted({t.split(" + ")[1] for t in production.loc[
    production["Plant Type"] == "Refinery", "Technology"]})
anodes = sorted({t.split(" + ")[0] for t in production.loc[
    production["Plant Type"] == "Smelter", "Technology"]})

abatement = []
for boiler in boilers:
    abatement.append(["Digester (refinery boiler)", boiler,
                      f"Scope 1 tCO2 per t alumina, boiler component, {EF_REGION} {EF_YEAR}",
                      ref_ef[f"{boiler} + Elec-Calciner"], "Gas-Boiler"])
for calciner in calciners:
    abatement.append(["Calciner (refinery)", calciner,
                      f"Scope 1 tCO2 per t alumina, calciner component, {EF_REGION} {EF_YEAR}",
                      ref_ef[f"Elec-Boiler + {calciner}"], "Gas-Calciner"])
for anode in anodes:
    abatement.append(["Anode (smelter)", anode,
                      f"Direct tCO2e per t aluminium, {ANODE_EF_REGION} {ANODE_EF_YEAR}, "
                      f"read on '+ Grid'", sm_ef[f"{anode} + Grid"], "Carbon Anode"])

# Expected milestone years, from derive_milestones.py, for the checks block.
EXPECTED = {("unabated fossil digester", "China"): (2021, 2042),
            ("unabated fossil digester", "RoW"): (2022, 2045),
            ("unabated fossil calciner", "China"): (2031, 2050),
            ("unabated fossil calciner", "RoW"): (2022, 2050),
            ("unabated carbon anode", "China"): (2021, 2042),
            ("unabated carbon anode", "RoW"): (2033, 2043),
            ("unabated fossil auxiliary", "China"): (2025, 2050),
            ("unabated fossil auxiliary", "RoW"): (2025, 2050)}


# ============================================================== workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ----------------------------------------------------------- Read Me
ws = wb.create_sheet("Read Me")
for row in [
    ["SBTi Aluminium Sector Pathway - Asset Transition Milestones"],
    [],
    ["Scenario", "MPP 1.5DS for digester, calciner and anode; IAI 1.5DS for auxiliary"],
    ["Regions", "China and Rest of the World (the finest split MPP publishes)"],
    ["Output", "Two years per technology group per region, in the Power specification format:"],
    ["", "a 'no new capacity' year and a 'phase out' year"],
    [],
    ["CAPTIVE POWER GENERATION IS OUT OF SCOPE"],
    ["", "MPP smelter technology strings are '<anode> + <power source>', where the power "
         "source is the"],
    ["", "smelter's captive or contracted electricity. Under SBTi guidance that generation "
         "takes its"],
    ["", "milestone from the power sector, from a different source, so it is never tested "
         "here. Anode"],
    ["", "milestones are computed on production summed across all power sources. Column G of "
         "Published"],
    ["", "Production shows the excluded power source so the exclusion is visible."],
    [],
    ["THE SEVEN MILESTONES"],
    ["1", "No new unabated fossil digester       refinery boiler raising digestion steam"],
    ["2", "Phase out unabated fossil digester"],
    ["3", "No new unabated fossil calciner       refinery calciner"],
    ["4", "Phase out unabated fossil calciner"],
    ["5", "Phase out unabated carbon anode       smelter anode"],
    ["6", "No new unabated fossil auxiliary      remelting, scrap refining, casting"],
    ["7", "Phase out unabated fossil auxiliary"],
    [],
    ["WHAT COUNTS AS UNABATED - see sheet Abatement Test"],
    ["", "A technology is ABATED if either limb holds:"],
    ["(a)", "it is fitted with CCS at 90% capture rate or better, or"],
    ["(b)", "its scope 1 intensity is 10% or less of BAU performance for the same process step."],
    ["", "Everything else is unabated. Both limbs are tested against data, not technology "
         "names,"],
    ["", "because MPP's names mislead in both directions - see the Abatement Test sheet."],
    [],
    ["THE TWO TESTS - see the three Milestone sheets"],
    ["No new", "The year after the last year production exceeded its own prior running maximum"],
    ["", "by more than the materiality threshold. Valid because MPP caps utilisation at "
         "CUF 0.95,"],
    ["", "so production above a technology's previous peak cannot come from utilisation alone "
         "and"],
    ["", "requires added capacity. Applied per technology and the latest year across the "
         "group"],
    ["", "taken, so building a new gas boiler while retiring a coal boiler still counts."],
    ["Phase out", "The first year the group total is at or below the materiality threshold and "
                  "stays there."],
    [],
    ["LIMITATION TO DISCLOSE"],
    ["", "The capacity test detects NET growth within a technology, so it is a lower bound on "
         "gross"],
    ["", "new build. Capacity added in a year while more capacity retired would be invisible."],
    [],
    ["SOURCES"],
    ["Production", "MPP published workbooks (1) China and (2) Rest of the World, sheet "
                   "Annual_production_volume_Mt_df"],
    ["Refinery factors", "mpp-shared-code repo input def_refineries/intermediate/emissions.csv"],
    ["Smelter factors", "mpp-shared-code repo input def/intermediate/emissions.csv"],
    ["Capture rate", "MPP Aluminium Technical Appendix p.11 - '90% capture of smelter process "
                     "CO2, not PFCs or anode production emissions'"],
    ["Auxiliary", "IAI 1.5 Scenario Dataset, sheet 1.5, Tables 2 and 3. MPP defers these value "
                  "chain steps to IAI (Appendix pp.3-4, 20)"],
    ["", "No local model run is involved anywhere in this workbook."],
]:
    ws.append(row)
ws["A1"].font = TITLE
for r in (8, 15, 24, 31, 39, 43):
    ws.cell(r, 1).font = HEAD
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 104

# ----------------------------------------------------------- Inputs
ws_in = wb.create_sheet("Inputs")
ws_in["A1"] = "Inputs - every formula in this workbook points at these cells"
ws_in["A1"].font = TITLE
inputs = [
    ("Parameter", "Value", "Named as", "Why this value"),
    ("Materiality threshold, share of 2020", 0.01, "MaterialityShare",
     "Without it, single stranded assets set the milestone. China's unabated carbon anode "
     "falls 99.7% by 2042, then one 0.13 Mt smelter holds flat to 2050."),
    ("Abated if intensity at or below", 0.10, "IntensityLimit",
     "Limb (b) of the abatement rule: roughly 10% of BAU performance."),
    ("Abated if CCS capture at or above", 0.90, "CaptureLimit",
     "Limb (a) of the abatement rule."),
    ("Net zero backstop year", 2050, "BackstopYear",
     "Net zero by 2050 is a hard requirement, so any milestone with no year in the source "
     "scenario is set to 2050 and flagged."),
    ("Equipment lifetime, years", 25, "EquipmentLifetime",
     "Auxiliary 'no new' year only. IAI publishes no lifetime; 25 years is MPP's own "
     "assumption for every boiler type (Appendix Exhibit TA3.1). CONFIRM BEFORE PUBLISHING."),
]
for row in inputs:
    ws_in.append(list(row))
for c in ws_in[2]:
    c.font = HEAD
for name, cell in [("MaterialityShare", "B3"), ("IntensityLimit", "B4"),
                   ("CaptureLimit", "B5"), ("BackstopYear", "B6"),
                   ("EquipmentLifetime", "B7")]:
    wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!${cell[0]}${cell[1:]}"))
for c, w in zip("ABCD", [36, 10, 20, 88]):
    ws_in.column_dimensions[c].width = w
for r in range(3, 8):
    ws_in.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="top")

# ----------------------------------------------------------- Abatement Test
ws_ab = wb.create_sheet("Abatement Test")
ws_ab["A1"] = "Abatement Test - which technologies are unabated, and why"
ws_ab["A1"].font = TITLE
ws_ab["A2"] = ("Component ratios are identical in every MPP region and year, so one "
               "region-year classifies the whole set.")
ws_ab.append([])
ws_ab.append(["Process Step", "Technology", "Metric", "Value", "Benchmark Technology",
              "Benchmark Value", "% of Benchmark", "CCS Capture Rate",
              "Abated by Capture?", "Abated by Intensity?", "Classification by Rule",
              "Override", "Reason for Override", "Classification Used"])
for c in ws_ab[4]:
    c.font = HEAD

AB_FIRST = 5
for i, (step, tech, metric, value, benchmark) in enumerate(abatement):
    r = AB_FIRST + i
    ws_ab.cell(r, 1, step)
    ws_ab.cell(r, 2, tech)
    ws_ab.cell(r, 3, metric)
    ws_ab.cell(r, 4, round(float(value), 6))
    ws_ab.cell(r, 5, benchmark)
    ws_ab.cell(r, 8, CAPTURE_RATES.get(tech, 0))
    override, reason = OVERRIDES.get(tech, ("", ""))
    ws_ab.cell(r, 12, override)
    ws_ab.cell(r, 13, reason)
AB_LAST = AB_FIRST + len(abatement) - 1

for r in range(AB_FIRST, AB_LAST + 1):
    ws_ab[f"F{r}"] = (f"=IFERROR(VLOOKUP($E{r},$B${AB_FIRST}:$D${AB_LAST},3,FALSE),\"\")")
    ws_ab[f"G{r}"] = f'=IF($F{r}="","",$D{r}/$F{r})'
    ws_ab[f"I{r}"] = f'=IF($H{r}>=CaptureLimit,"Abated","")'
    ws_ab[f"J{r}"] = f'=IF(AND($G{r}<>"",$G{r}<=IntensityLimit),"Abated","")'
    ws_ab[f"K{r}"] = f'=IF(OR($I{r}="Abated",$J{r}="Abated"),"Abated","Unabated")'
    ws_ab[f"N{r}"] = f'=IF($L{r}<>"",$L{r},$K{r})'
    ws_ab[f"G{r}"].number_format = "0.0%"
    ws_ab[f"H{r}"].number_format = "0%"

for c, w in zip("ABCDEFGHIJKLMN",
                [26, 20, 56, 10, 20, 14, 14, 12, 14, 16, 20, 12, 60, 20]):
    ws_ab.column_dimensions[c].width = w
for r in range(AB_FIRST, AB_LAST + 1):
    ws_ab.cell(r, 13).alignment = Alignment(wrap_text=True, vertical="top")

AB = "'Abatement Test'"
AB_LOOKUP = f"{AB}!$B${AB_FIRST}:$N${AB_LAST}"

# ----------------------------------------------------------- Published Production
ws_pp = wb.create_sheet("Published Production")
ws_pp.append(["Region", "Plant Type", "Technology", "Boiler", "Calciner", "Anode",
              "Power Source - EXCLUDED FROM SCOPE", "Year", "Production Mt",
              "Digester Classification", "Calciner Classification", "Anode Classification"])
for c in ws_pp[1]:
    c.font = HEAD
for region, plant_type, tech, year, value in production.itertuples(index=False, name=None):
    ws_pp.append([region, plant_type, tech, None, None, None, None,
                  int(year), float(value)])
PP_LAST = len(production) + 1

for r in range(2, PP_LAST + 1):
    ws_pp[f"D{r}"] = f'=IF($B{r}="Refinery",LEFT($C{r},FIND(" + ",$C{r})-1),"")'
    ws_pp[f"E{r}"] = f'=IF($B{r}="Refinery",MID($C{r},FIND(" + ",$C{r})+3,100),"")'
    ws_pp[f"F{r}"] = f'=IF($B{r}="Smelter",LEFT($C{r},FIND(" + ",$C{r})-1),"")'
    ws_pp[f"G{r}"] = f'=IF($B{r}="Smelter",MID($C{r},FIND(" + ",$C{r})+3,100),"")'
    ws_pp[f"J{r}"] = f'=IFERROR(VLOOKUP($D{r},{AB_LOOKUP},13,FALSE),"")'
    ws_pp[f"K{r}"] = f'=IFERROR(VLOOKUP($E{r},{AB_LOOKUP},13,FALSE),"")'
    ws_pp[f"L{r}"] = f'=IFERROR(VLOOKUP($F{r},{AB_LOOKUP},13,FALSE),"")'
for c, w in zip("ABCDEFGHIJKL",
                [10, 12, 34, 20, 16, 18, 34, 8, 14, 22, 22, 22]):
    ws_pp.column_dimensions[c].width = w
ws_pp.freeze_panes = "A2"

PP = "'Published Production'"
REG = f"{PP}!$A$2:$A${PP_LAST}"
YR = f"{PP}!$H$2:$H${PP_LAST}"
VAL = f"{PP}!$I$2:$I${PP_LAST}"


# ----------------------------------------------------------- Milestone sheets
def build_milestone_sheet(spec):
    """One sheet per technology group, one vertical block per region."""
    ws_m = wb.create_sheet(spec["sheet"])
    ws_m["A1"] = f"Milestones - {spec['label']}"
    ws_m["A1"].font = TITLE
    ws_m["A2"] = (f"Production in {spec['unit']}, summed across every published technology "
                  f"whose component is classified Unabated on the Abatement Test sheet.")

    techs = spec["technologies"]
    comp = f"{PP}!${spec['component_col']}$2:${spec['component_col']}${PP_LAST}"
    cls = f"{PP}!${spec['class_col']}$2:${spec['class_col']}${PP_LAST}"

    results = {}
    top = 4
    for region in REGIONS:
        key = top + 3                       # bare technology names, matched by SUMIFS
        head = key + 1                      # column header row
        first = head + 1                    # first year row
        last = first + len(YEARS) - 1
        thresh = f"$B${top + 1}"

        ws_m.cell(top, 1, region).font = SUBHEAD
        ws_m.cell(top + 1, 1, "Materiality threshold").font = HEAD
        ws_m.cell(top + 2, 1, "2020 group total x MaterialityShare")

        # The key row is what the SUMIFS below match on, so the technology name is visible
        # rather than buried in a formula.
        ws_m.cell(key, 1, "Technology matched ->").font = HEAD
        for i, tech in enumerate(techs):
            ws_m.cell(key, 2 + 4 * i, tech).font = HEAD

        # Column headers: four per technology, then the group block.
        headers = ["Year"]
        for tech in techs:
            headers += [f"{tech} Production", f"{tech} Prior Running Max",
                        f"{tech} Capacity Added?", f"{tech} Addition Year"]
        headers += ["Group Total (Unabated)", "Sum of Columns Above (check)",
                    "Max This Year Onward", "At or Below Threshold Thereafter?",
                    "Phase Out Candidate Year"]
        for j, text in enumerate(headers, start=1):
            c = ws_m.cell(head, j, text)
            c.font = HEAD
            c.alignment = Alignment(wrap_text=True, vertical="bottom")

        total_col = get_column_letter(2 + 4 * len(techs))
        check_col = get_column_letter(3 + 4 * len(techs))
        onward_col = get_column_letter(4 + 4 * len(techs))
        stays_col = get_column_letter(5 + 4 * len(techs))
        cand_col = get_column_letter(6 + 4 * len(techs))

        ws_m.cell(top + 1, 2).value = f"={total_col}{first}*MaterialityShare"
        ws_m.cell(top + 1, 2).number_format = "0.000"

        for k, year in enumerate(YEARS):
            r = first + k
            ws_m.cell(r, 1, int(year))
            for i, tech in enumerate(techs):
                p = get_column_letter(2 + 4 * i)
                m = get_column_letter(3 + 4 * i)
                a = get_column_letter(4 + 4 * i)
                y = get_column_letter(5 + 4 * i)
                ws_m[f"{p}{r}"] = (f"=SUMIFS({VAL},{REG},$A${top},{comp},{p}${key},"
                                   f"{YR},$A{r})")
                if k == 0:
                    ws_m[f"{m}{r}"] = "n/a - initial stock"
                    ws_m[f"{a}{r}"] = "n/a"
                else:
                    ws_m[f"{m}{r}"] = f"=MAX({p}${first}:{p}{r - 1})"
                    ws_m[f"{a}{r}"] = (f'=IF({p}{r}>{m}{r}+{thresh},"Yes","No")')
                ws_m[f"{y}{r}"] = f'=IF({a}{r}="Yes",$A{r},"")'
                ws_m[f"{p}{r}"].number_format = "0.00"
                ws_m[f"{m}{r}"].number_format = "0.00"

            ws_m[f"{total_col}{r}"] = (f'=SUMIFS({VAL},{REG},$A${top},{cls},"Unabated",'
                                       f"{YR},$A{r})")
            ws_m[f"{check_col}{r}"] = "=" + "+".join(
                f"{get_column_letter(2 + 4 * i)}{r}" for i in range(len(techs)))
            ws_m[f"{onward_col}{r}"] = f"=MAX({total_col}{r}:{total_col}${last})"
            ws_m[f"{stays_col}{r}"] = (f'=IF({onward_col}{r}<={thresh},"Yes","No")')
            ws_m[f"{cand_col}{r}"] = f'=IF({stays_col}{r}="Yes",$A{r},"")'
            ws_m[f"{total_col}{r}"].number_format = "0.00"
            ws_m[f"{check_col}{r}"].number_format = "0.00"
            ws_m[f"{onward_col}{r}"].number_format = "0.00"

        # Results block
        res = last + 2
        addition_ranges = ",".join(
            f"{get_column_letter(5 + 4 * i)}${first}:{get_column_letter(5 + 4 * i)}${last}"
            for i in range(len(techs)))
        candidates = f"{cand_col}${first}:{cand_col}${last}"

        ws_m.cell(res, 1, f"RESULTS - {region}").font = SUBHEAD
        rows = [
            ("Last addition year, any unabated technology",
             f'=IF(COUNT({addition_ranges})=0,"none",MAX({addition_ranges}))'),
            ("NO NEW YEAR",
             f"=IF(COUNT({addition_ranges})=0,2021,MAX({addition_ranges})+1)"),
            ("Phase out year in the source scenario",
             f'=IF(COUNT({candidates})=0,"beyond 2050",MIN({candidates}))'),
            ("PHASE OUT YEAR",
             f"=IF(COUNT({candidates})=0,BackstopYear,MIN({candidates}))"),
            ("Backstop applied?",
             f'=IF(COUNT({candidates})=0,"Yes","No")'),
            ("Group total check - all years agree?",
             f'=IF(SUMPRODUCT(--(ROUND({total_col}${first}:{total_col}${last},6)'
             f'<>ROUND({check_col}${first}:{check_col}${last},6)))=0,"PASS","FAIL")'),
        ]
        for j, (label, formula) in enumerate(rows, start=1):
            ws_m.cell(res + j, 1, label)
            ws_m.cell(res + j, 2).value = formula
        ws_m.cell(res + 2, 1).font = HEAD
        ws_m.cell(res + 2, 2).font = HEAD
        ws_m.cell(res + 4, 1).font = HEAD
        ws_m.cell(res + 4, 2).font = HEAD

        results[region] = {"no_new": f"'{spec['sheet']}'!$B${res + 2}",
                           "phase_out": f"'{spec['sheet']}'!$B${res + 4}",
                           "backstop": f"'{spec['sheet']}'!$B${res + 5}"}
        top = res + 9

    ws_m.column_dimensions["A"].width = 40
    for j in range(2, 3 + 4 * len(techs) + 5):
        ws_m.column_dimensions[get_column_letter(j)].width = 16
    return results


milestone_refs = {}
for spec in MILESTONES:
    milestone_refs[spec["label"]] = build_milestone_sheet(spec)

# ----------------------------------------------------------- Auxiliary IAI
iai = pd.read_excel(DATA_DIR / "IAI_1.5 Scenario Dataset.xlsx", sheet_name="1.5",
                    header=None)


def iai_table(marker):
    start = iai.index[iai[0].astype(str).str.startswith(marker)][0]
    years = iai.iloc[start + 3, 1:7].astype(int).tolist()
    table = {}
    for i in range(start + 4, len(iai)):
        label = iai.iloc[i, 0]
        if pd.isna(label):
            break
        table[label] = pd.Series(iai.iloc[i, 1:7].astype(float).values, index=years)
    return pd.DataFrame(table).T


iai_emissions = iai_table("Table 2")
iai_production = iai_table("Table 3")
IAI_YEARS = list(iai_emissions.columns)

ws_ax = wb.create_sheet("Auxiliary IAI")
ws_ax["A1"] = "Milestones - unabated fossil auxiliary (remelting, scrap refining, casting)"
ws_ax["A1"].font = TITLE
ws_ax["A2"] = ("IAI 1.5DS, GLOBAL ONLY - no regional split, so the same year is reported for "
               "China and RoW.")
ws_ax["A3"] = ("IAI has no technology mix or capacity dimension, so the production-growth "
               "test cannot be used. Phase out is read against the 10% intensity limb "
               "directly; no new is the phase-out year less the equipment lifetime.")
ws_ax["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws_ax.row_dimensions[3].height = 30

AX_HEAD = 5
ws_ax.cell(AX_HEAD, 1, "IAI 1.5 Scenario Dataset, sheet 1.5").font = HEAD
for j, year in enumerate(IAI_YEARS, start=2):
    ws_ax.cell(AX_HEAD, j, int(year)).font = HEAD

r = AX_HEAD
for step in IAI_STEPS:
    r += 1
    ws_ax.cell(r, 1, f"Table 2 emissions, {step} (Mt CO2e)")
    for j, year in enumerate(IAI_YEARS, start=2):
        ws_ax.cell(r, j, round(float(iai_emissions.loc[step, year]), 4))
STEP_FIRST, STEP_LAST = AX_HEAD + 1, r

r += 1
TOTAL_ROW = r
ws_ax.cell(r, 1, "Total auxiliary emissions (Mt CO2e)").font = HEAD
r += 1
SEMIS_ROW = r
ws_ax.cell(r, 1, f"Table 3 production, {IAI_DENOMINATOR} (Mt)")
for j, year in enumerate(IAI_YEARS, start=2):
    ws_ax.cell(r, j, round(float(iai_production.loc[IAI_DENOMINATOR, year]), 4))
r += 2
INT_ROW = r
ws_ax.cell(r, 1, "Auxiliary intensity (tCO2e per t semis)").font = HEAD
r += 1
PCT_ROW = r
ws_ax.cell(r, 1, f"% of {IAI_YEARS[0]}").font = HEAD
r += 1
ABATED_ROW = r
ws_ax.cell(r, 1, "At or below the intensity limit?")
r += 1
CAND_ROW = r
ws_ax.cell(r, 1, "Phase out candidate year")

for j, year in enumerate(IAI_YEARS, start=2):
    col = get_column_letter(j)
    ws_ax.cell(TOTAL_ROW, j).value = f"=SUM({col}{STEP_FIRST}:{col}{STEP_LAST})"
    ws_ax.cell(INT_ROW, j).value = f"={col}{TOTAL_ROW}/{col}{SEMIS_ROW}"
    ws_ax.cell(PCT_ROW, j).value = f"={col}{INT_ROW}/$B${INT_ROW}"
    ws_ax.cell(ABATED_ROW, j).value = f'=IF({col}{PCT_ROW}<=IntensityLimit,"Yes","No")'
    ws_ax.cell(CAND_ROW, j).value = f'=IF({col}{ABATED_ROW}="Yes",{col}${AX_HEAD},"")'
    ws_ax.cell(TOTAL_ROW, j).number_format = "0.00"
    ws_ax.cell(INT_ROW, j).number_format = "0.000"
    ws_ax.cell(PCT_ROW, j).number_format = "0.0%"

LAST_COL = get_column_letter(1 + len(IAI_YEARS))
CANDS = f"$B${CAND_ROW}:${LAST_COL}${CAND_ROW}"
res = CAND_ROW + 2
ws_ax.cell(res, 1, "RESULTS - global, applied to both regions").font = SUBHEAD
for j, (label, formula) in enumerate([
    ("Phase out year in the source scenario",
     f'=IF(COUNT({CANDS})=0,"beyond 2050",MIN({CANDS}))'),
    ("PHASE OUT YEAR", f"=IF(COUNT({CANDS})=0,BackstopYear,MIN({CANDS}))"),
    ("Backstop applied?", f'=IF(COUNT({CANDS})=0,"Yes","No")'),
    ("NO NEW YEAR", f"=$B${res + 2}-EquipmentLifetime"),
    ("", "Phase out year less the equipment lifetime, because a fossil furnace built in "
         "year Y is still operating in Y + lifetime"),
], start=1):
    ws_ax.cell(res + j, 1, label)
    ws_ax.cell(res + j, 2).value = formula
ws_ax.cell(res + 2, 1).font = HEAD
ws_ax.cell(res + 2, 2).font = HEAD
ws_ax.cell(res + 4, 1).font = HEAD
ws_ax.cell(res + 4, 2).font = HEAD

ws_ax.cell(res + 7, 1, "ASSUMPTIONS TO CONFIRM BEFORE PUBLISHING").font = HEAD
for j, text in enumerate([
    "25-year equipment lifetime. IAI publishes no lifetime; this is MPP's boiler assumption "
    "(Appendix Exhibit TA3.1). A 20-year lifetime would give 2030 instead of 2025.",
    "Primary casting is not covered. IAI folds it into Primary Aluminium as an unlabelled "
    "residual - 2018: 1036.6 total less 823.3 electrolysis less 171.5 refining = 41.8 Mt, "
    "which also carries bauxite mining and anode production.",
], start=1):
    ws_ax.cell(res + 7 + j, 1, text).alignment = Alignment(wrap_text=True, vertical="top")
    ws_ax.row_dimensions[res + 7 + j].height = 28

ws_ax.column_dimensions["A"].width = 52
for j in range(2, 2 + len(IAI_YEARS)):
    ws_ax.column_dimensions[get_column_letter(j)].width = 13

AX = "'Auxiliary IAI'"
milestone_refs["unabated fossil auxiliary"] = {
    region: {"no_new": f"{AX}!$B${res + 4}", "phase_out": f"{AX}!$B${res + 2}",
             "backstop": f"{AX}!$B${res + 3}"} for region in REGIONS}

# ----------------------------------------------------------- Milestone Summary
ws_sum = wb.create_sheet("Milestone Summary")
ws_sum["A1"] = "SBTi Aluminium Sector Pathway - Asset Transition Milestones"
ws_sum["A1"].font = TITLE
ws_sum["A2"] = "MPP 1.5DS; IAI 1.5DS for auxiliary. Captive power generation out of scope."

SPEC_ROWS = [
    ("No new unabated fossil digester", "unabated fossil digester", "no_new"),
    ("Phase out unabated fossil digester", "unabated fossil digester", "phase_out"),
    ("No new unabated fossil calciner", "unabated fossil calciner", "no_new"),
    ("Phase out unabated fossil calciner", "unabated fossil calciner", "phase_out"),
    ("Phase out unabated carbon anode", "unabated carbon anode", "phase_out"),
    ("No new unabated fossil auxiliary", "unabated fossil auxiliary", "no_new"),
    ("Phase out unabated fossil auxiliary", "unabated fossil auxiliary", "phase_out"),
]

ws_sum.append([])
ws_sum.append(["Milestone", "China", "RoW", "China backstop?", "RoW backstop?"])
for c in ws_sum[4]:
    c.font = HEAD
for i, (label, group, kind) in enumerate(SPEC_ROWS):
    r = 5 + i
    ws_sum.cell(r, 1, label)
    for j, region in enumerate(REGIONS, start=2):
        ws_sum.cell(r, j).value = f"={milestone_refs[group][region][kind]}"
        flag = "" if kind == "no_new" else f"={milestone_refs[group][region]['backstop']}"
        ws_sum.cell(r, j + 2).value = flag
SPEC_LAST = 4 + len(SPEC_ROWS)

note = SPEC_LAST + 2
ws_sum.cell(note, 1, "Where 'backstop?' is Yes, the year comes from the 2050 net-zero "
                     "requirement, not from the source scenario. Disclose these.")
ws_sum.cell(note + 1, 1, "Auxiliary carries the same year in both regions - IAI 1.5DS is "
                         "global, with no regional disaggregation.")
ws_sum.cell(note + 2, 1, "'Phase out unabated carbon anode' does NOT imply zero smelter "
                         "process emissions: HH+CCS is abated on the 90% capture limb but "
                         "leaves about 1.03 tCO2e/tAl of PFCs and anode production emissions.")

chk = note + 4
ws_sum.cell(chk, 1, "CHECKS - workbook formulas against derive_milestones.py").font = SUBHEAD
ws_sum.cell(chk + 1, 1, "Milestone").font = HEAD
for j, text in enumerate(["Region", "Expected No New", "Workbook No New", "No New",
                          "Expected Phase Out", "Workbook Phase Out", "Phase Out"],
                         start=2):
    ws_sum.cell(chk + 1, j, text).font = HEAD

r = chk + 1
for group in ["unabated fossil digester", "unabated fossil calciner",
              "unabated carbon anode", "unabated fossil auxiliary"]:
    for region in REGIONS:
        r += 1
        expected_no_new, expected_phase_out = EXPECTED[(group, region)]
        ws_sum.cell(r, 1, group)
        ws_sum.cell(r, 2, region)
        ws_sum.cell(r, 3, expected_no_new)
        ws_sum.cell(r, 4).value = f"={milestone_refs[group][region]['no_new']}"
        ws_sum.cell(r, 5).value = f'=IF($C{r}=$D{r},"PASS","FAIL")'
        ws_sum.cell(r, 6, expected_phase_out)
        ws_sum.cell(r, 7).value = f"={milestone_refs[group][region]['phase_out']}"
        ws_sum.cell(r, 8).value = f'=IF($F{r}=$G{r},"PASS","FAIL")'

for c, w in zip("ABCDEFGH", [40, 14, 16, 16, 10, 18, 18, 10]):
    ws_sum.column_dimensions[c].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

print(f"wrote {OUT}")
print(f"  Abatement Test        {len(abatement):5d} technologies")
print(f"  Published Production  {len(production):5d} rows "
      f"({len(REGIONS)} regions x {production['Technology'].nunique()} technologies "
      f"x {len(YEARS)} years)")
print(f"  Milestone sheets      {len(MILESTONES):5d} + Auxiliary IAI")
print(f"  Checks                {len(EXPECTED) * 2:5d} comparisons against derive_milestones.py")
