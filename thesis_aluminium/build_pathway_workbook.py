"""Build the aluminium process-intensity pathway workbook from MPP published data.

Everything here comes from MPP artefacts only — no local model run:
  - production by technology x year x {China, Rest of the World} from MPP's published
    workbooks (the two regional files are an exact partition of the global one)
  - anode archetype constants from MPP Aluminium Technical Appendix, Exhibit TA3.3
  - refinery emission factors from the unmodified repo input emissions.csv
  - alumina-per-aluminium ratio from the unmodified repo input inputs_outputs.csv

The workbook is formula-driven end to end so the derivation can be traced in Excel.
Only the raw source tables are written as values.
"""

from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = Path(".")
MPP = DATA_DIR / "mpp-shared-code" / "aluminium" / "data" / "lc"
OUT = Path("/Users/parkercaswell/Desktop/Aluminium Pathway Derivation/Aluminium Emissions Pathway.xlsx")

PUBLISHED = {"China": "mpp_aluminium_net_zero_outputs (1).xlsx",
             "RoW": "mpp_aluminium_net_zero_outputs (2).xlsx"}
SCENARIO = "1.5DS"

# Anode archetype direct-emission constants, tCO2e per t Al (Tech Appendix Exhibit TA3.3).
# Longest prefix first so Carbon Anode+CCS is matched before Carbon Anode.
ANODE = [("Carbon Anode+CCS", 1.17), ("Carbon Anode", 2.13), ("Inert Anode", 0.10)]
OUT_YEARS = [2020, 2025, 2030, 2035, 2040, 2045, 2050]


def bold(ws, row=1):
    for c in ws[row]:
        c.font = Font(bold=True)


def write(wb, name, df):
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    bold(ws)
    return ws


def load_repo(path):
    df = pd.read_csv(path)
    df.columns = [str(c).strip().lstrip("﻿") for c in df.columns]
    return df


def anode_constant(tech):
    for prefix, value in ANODE:
        if tech.startswith(prefix):
            return value
    raise ValueError(f"unmapped smelter technology: {tech}")


def archetype(tech):
    for prefix, _ in ANODE:
        if tech.startswith(prefix):
            return prefix
    raise ValueError(tech)


# ---------------------------------------------------------------- source data
published = []
for group, filename in PUBLISHED.items():
    df = pd.read_excel(DATA_DIR / filename, sheet_name="Annual_production_volume_Mt_df")
    df = df[df["scenario"] == SCENARIO].copy()
    df["Region Group"] = group
    published.append(df)
published = pd.concat(published)

smelter = published[published["plant_type"] == "Smelter"][
    ["Region Group", "technology", "year", "value"]]
smelter.columns = ["Region Group", "Technology", "Year", "Production Mt Al"]
smelter = smelter.sort_values(["Region Group", "Technology", "Year"]).reset_index(drop=True)

refinery = published[published["plant_type"] == "Refinery"][
    ["Region Group", "technology", "year", "value"]]
refinery.columns = ["Region Group", "Technology", "Year", "Production Mt Alumina"]
refinery = refinery.sort_values(["Region Group", "Technology", "Year"]).reset_index(drop=True)

# Refinery emission factors, by MPP model region. China sub-regions are identical for a
# given technology-year, so China needs no weighting; RoW does.
ref_ef = load_repo(MPP / "def_refineries" / "intermediate" / "emissions.csv")
ref_ef["Region Group"] = ["China" if str(r).startswith("China") else "RoW"
                          for r in ref_ef["region"]]
ref_ef = ref_ef[["region", "Region Group", "technology", "year", "co2_scope1"]]
ref_ef.columns = ["MPP Region", "Region Group", "Technology", "Year",
                  "Emission Factor tCO2e per t Alumina"]
ref_ef = ref_ef.sort_values(["Region Group", "Technology", "Year", "MPP Region"]).reset_index(drop=True)

# Weights for collapsing MPP regions to China / RoW: 2020 refinery capacity share.
init = load_repo(MPP / "def_refineries" / "intermediate" / "initial_asset_stack.csv")
init["Region Group"] = ["China" if str(r).startswith("China") else "RoW"
                        for r in init["region"]]
weights = (init.groupby(["Region Group", "region"])["annual_production_capacity"]
           .sum().reset_index())
weights.columns = ["Region Group", "MPP Region", "Capacity 2020 Mt Alumina"]

# Alumina per aluminium, read from data rather than assumed.
io = load_repo(MPP / "def" / "intermediate" / "inputs_outputs.csv")
alumina_ratio = io.loc[io["parameter"] == "Alumina Consumption", "value"].unique()
assert len(alumina_ratio) == 1, f"alumina ratio is not a single constant: {alumina_ratio}"
ALUMINA_RATIO = float(alumina_ratio[0])

anode_table = pd.DataFrame({
    "Technology": sorted(smelter["Technology"].unique())})
anode_table["Anode Archetype"] = anode_table["Technology"].map(archetype)
anode_table["Constant tCO2e per t Al"] = anode_table["Technology"].map(anode_constant)

# ---------------------------------------------------------------- build workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- Read Me
ws = wb.create_sheet("Read Me")
for row in [
    ["SBTi Aluminium Sector Pathway — Process Emissions Intensity"],
    [],
    ["Scenario", "MPP 1.5DS"],
    ["Metric", "tCO2e per tonne aluminium"],
    ["Regions", "China and Rest of the World (the finest split MPP publishes)"],
    ["Boundary", "Process emissions ONLY — excludes all electricity, both Scope 2 imported "
                 "power and captive fossil generation"],
    ["In scope", "Smelter anode process CO2 + PFCs + anode thermal; alumina refining "
                 "digestion and calcination fuel"],
    [],
    ["SOURCES — all MPP, no local model run"],
    ["Smelter and refinery production", "MPP published workbooks (1) China and (2) Rest of the World"],
    ["Anode constants", "MPP Aluminium Technical Appendix, Exhibit TA3.3 p.11"],
    ["Refinery emission factors", "mpp-shared-code repo input def_refineries/intermediate/emissions.csv"],
    ["Alumina per aluminium", f"{ALUMINA_RATIO} t/t, repo input inputs_outputs.csv (single global constant)"],
    [],
    ["METHOD"],
    ["1", "Smelter process intensity = SUM(production x anode constant) / SUM(production)."],
    ["", "Anode constants are region-independent, so no weighting is needed."],
    ["2", "Refinery emission factors are published per MPP model region. They are collapsed to"],
    ["", "China / RoW by 2020 refinery capacity share. China sub-regions are identical for a"],
    ["", "given technology-year so China is exact; RoW is a weighted average."],
    ["3", "Refinery intensity per t alumina = SUM(production x factor) / SUM(production)."],
    ["4", f"Converted to a per-tonne-aluminium basis by multiplying by {ALUMINA_RATIO}."],
    ["5", "Process intensity = smelter intensity + refinery intensity per t Al."],
    [],
    ["KNOWN APPROXIMATION"],
    ["", "RoW refinery factors vary across MPP regions within a technology-year and published"],
    ["", "data does not resolve production below RoW. Capacity weighting versus equal weighting"],
    ["", "differs by 0.10 tCO2e/tAl in 2020, 0.03 by 2040, and is negligible by 2050."],
]:
    ws.append(row)
ws["A1"].font = Font(bold=True, size=14)
for r in (9, 15, 25):
    ws.cell(r, 1).font = Font(bold=True)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 95

# --- source tables
write(wb, "Anode Constants", anode_table)
write(wb, "Region Weights", weights)

ws_sm = write(wb, "Smelter Production", smelter)
ws_rf = write(wb, "Refinery Production", refinery)
ws_raw = write(wb, "Refinery EF by Region", ref_ef)

N_SM, N_RF, N_RAW = len(smelter) + 1, len(refinery) + 1, len(ref_ef) + 1
N_ANODE = len(anode_table) + 1
N_W = len(weights) + 1

# --- Refinery EF by Region: attach the region weight and the weighted product
ws_raw["F1"], ws_raw["G1"], ws_raw["H1"] = "Region Weight", "Factor x Weight", "Lookup Key"
bold(ws_raw)
for r in range(2, N_RAW + 1):
    ws_raw[f"F{r}"] = f"=SUMIFS('Region Weights'!$C$2:$C${N_W},'Region Weights'!$B$2:$B${N_W},$A{r})"
    ws_raw[f"G{r}"] = f"=E{r}*F{r}"
    ws_raw[f"H{r}"] = f'=$B{r}&"|"&$C{r}&"|"&$D{r}'

# --- Refinery Emission Factors: capacity-weighted average per group-technology-year
combos = ref_ef[["Region Group", "Technology", "Year"]].drop_duplicates().sort_values(
    ["Region Group", "Technology", "Year"]).reset_index(drop=True)
ws_ef = write(wb, "Refinery Emission Factors", combos)
ws_ef["D1"], ws_ef["E1"] = "Weighted Factor tCO2e per t Alumina", "Lookup Key"
bold(ws_ef)
RAW = "'Refinery EF by Region'"
for r in range(2, len(combos) + 2):
    ws_ef[f"D{r}"] = (
        f"=SUMIFS({RAW}!$G$2:$G${N_RAW},{RAW}!$B$2:$B${N_RAW},$A{r},"
        f"{RAW}!$C$2:$C${N_RAW},$B{r},{RAW}!$D$2:$D${N_RAW},$C{r})"
        f"/SUMIFS({RAW}!$F$2:$F${N_RAW},{RAW}!$B$2:$B${N_RAW},$A{r},"
        f"{RAW}!$C$2:$C${N_RAW},$B{r},{RAW}!$D$2:$D${N_RAW},$C{r})")
    ws_ef[f"E{r}"] = f'=$A{r}&"|"&$B{r}&"|"&$C{r}'
N_EF = len(combos) + 1

# --- Smelter Production: anode constant and process emissions
ws_sm["E1"], ws_sm["F1"] = "Anode Constant tCO2e per t Al", "Process Emissions Mt"
bold(ws_sm)
for r in range(2, N_SM + 1):
    ws_sm[f"E{r}"] = f"=VLOOKUP($B{r},'Anode Constants'!$A$2:$C${N_ANODE},3,FALSE)"
    ws_sm[f"F{r}"] = f"=D{r}*E{r}"

# --- Refinery Production: emission factor and process emissions
ws_rf["E1"], ws_rf["F1"] = "Emission Factor tCO2e per t Alumina", "Process Emissions Mt"
bold(ws_rf)
EF = "'Refinery Emission Factors'"
for r in range(2, N_RF + 1):
    ws_rf[f"E{r}"] = (f'=INDEX({EF}!$D$2:$D${N_EF},MATCH($A{r}&"|"&$B{r}&"|"&$C{r},'
                      f"{EF}!$E$2:$E${N_EF},0))")
    ws_rf[f"F{r}"] = f"=D{r}*E{r}"

# --- Pathway Calculation
groups = ["China", "RoW", "Global"]
years = sorted(smelter["Year"].unique())
ws_calc = wb.create_sheet("Pathway Calculation")
ws_calc.append(["Region Group", "Year", "Production Mt Al", "Smelter Process Emissions Mt",
                "Production Mt Alumina", "Refinery Process Emissions Mt",
                "Smelter Intensity tCO2e per t Al", "Refinery Intensity tCO2e per t Al",
                "Process Intensity tCO2e per t Al", "Lookup Key"])
bold(ws_calc)
ws_calc["L1"] = "Alumina per Aluminium t/t"
ws_calc["L2"] = ALUMINA_RATIO
ws_calc["M1"] = "Source"
ws_calc["M2"] = ("repo input inputs_outputs.csv, parameter 'Alumina Consumption' — "
                 "single global constant, zero variation")
ws_calc["L1"].font = ws_calc["M1"].font = Font(bold=True)

SM, RF = "'Smelter Production'", "'Refinery Production'"
r = 1
for grp in groups:
    for y in years:
        r += 1
        ws_calc[f"A{r}"], ws_calc[f"B{r}"] = grp, int(y)
        if grp == "Global":
            ws_calc[f"C{r}"] = f"=SUMIFS({SM}!$D$2:$D${N_SM},{SM}!$C$2:$C${N_SM},$B{r})"
            ws_calc[f"D{r}"] = f"=SUMIFS({SM}!$F$2:$F${N_SM},{SM}!$C$2:$C${N_SM},$B{r})"
            ws_calc[f"E{r}"] = f"=SUMIFS({RF}!$D$2:$D${N_RF},{RF}!$C$2:$C${N_RF},$B{r})"
            ws_calc[f"F{r}"] = f"=SUMIFS({RF}!$F$2:$F${N_RF},{RF}!$C$2:$C${N_RF},$B{r})"
        else:
            ws_calc[f"C{r}"] = (f"=SUMIFS({SM}!$D$2:$D${N_SM},{SM}!$A$2:$A${N_SM},$A{r},"
                                f"{SM}!$C$2:$C${N_SM},$B{r})")
            ws_calc[f"D{r}"] = (f"=SUMIFS({SM}!$F$2:$F${N_SM},{SM}!$A$2:$A${N_SM},$A{r},"
                                f"{SM}!$C$2:$C${N_SM},$B{r})")
            ws_calc[f"E{r}"] = (f"=SUMIFS({RF}!$D$2:$D${N_RF},{RF}!$A$2:$A${N_RF},$A{r},"
                                f"{RF}!$C$2:$C${N_RF},$B{r})")
            ws_calc[f"F{r}"] = (f"=SUMIFS({RF}!$F$2:$F${N_RF},{RF}!$A$2:$A${N_RF},$A{r},"
                                f"{RF}!$C$2:$C${N_RF},$B{r})")
        ws_calc[f"G{r}"] = f"=D{r}/C{r}"
        ws_calc[f"H{r}"] = f"=F{r}/E{r}*$L$2"
        ws_calc[f"I{r}"] = f"=G{r}+H{r}"
        ws_calc[f"J{r}"] = f'=$A{r}&"|"&$B{r}'
N_CALC = r
CALC = "'Pathway Calculation'"

# --- Pathway Output
ws_out = wb.create_sheet("Pathway Output")
ws_out["A1"] = "SBTi Aluminium Sector Pathway — Process Emissions Intensity"
ws_out["A1"].font = Font(bold=True, size=13)
ws_out["A2"] = "MPP 1.5DS. Process emissions only — excludes all electricity."
ws_out["A3"] = "Built entirely from MPP published outputs and MPP repo inputs. No local model run."

blocks = [
    ("Process Intensity tCO2e per t Al", "I", 5),
    ("Reduction vs 2020 (%)", None, 16),
    ("Production Mt Al", "C", 27),
    ("Absolute Process Emissions Mt CO2e", None, 38),
]
for title, col, top in blocks:
    ws_out.cell(top, 1, title).font = Font(bold=True)
    ws_out.cell(top + 1, 1, "Year").font = Font(bold=True)
    for i, g in enumerate(groups):
        ws_out.cell(top + 1, 2 + i, g).font = Font(bold=True)
    for k, y in enumerate(OUT_YEARS):
        row = top + 2 + k
        ws_out.cell(row, 1, y)
        for i, g in enumerate(groups):
            if title.startswith("Reduction"):
                c = chr(ord("B") + i)
                ws_out.cell(row, 2 + i).value = f"=(1-{c}{7 + k}/{c}$7)*100"
            elif title.startswith("Absolute"):
                ws_out.cell(row, 2 + i).value = (
                    f'=INDEX({CALC}!$D$2:$D${N_CALC},MATCH("{g}|"&$A{row},{CALC}!$J$2:$J${N_CALC},0))'
                    f'+INDEX({CALC}!$F$2:$F${N_CALC},MATCH("{g}|"&$A{row},{CALC}!$J$2:$J${N_CALC},0))')
            else:
                ws_out.cell(row, 2 + i).value = (
                    f'=INDEX({CALC}!${col}$2:${col}${N_CALC},'
                    f'MATCH("{g}|"&$A{row},{CALC}!$J$2:$J${N_CALC},0))')

for c, w in zip("ABCDE", [38, 12, 12, 12, 12]):
    ws_out.column_dimensions[c].width = w
for c, w in zip("ABCDEFGHIJ", [14, 8, 18, 22, 20, 24, 22, 22, 22, 16]):
    ws_calc.column_dimensions[c].width = w

wb.save(OUT)
print(f"wrote {OUT}")
print(f"  Smelter Production        {len(smelter):6d} rows")
print(f"  Refinery Production       {len(refinery):6d} rows")
print(f"  Refinery EF by Region     {len(ref_ef):6d} rows")
print(f"  Refinery Emission Factors {len(combos):6d} rows")
print(f"  Pathway Calculation       {N_CALC - 1:6d} rows  ({len(groups)} groups x {len(years)} years)")
print(f"  alumina ratio             {ALUMINA_RATIO}")
